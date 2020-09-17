# -*- coding: utf-8 -*-
from typing import Dict, Any

import requests
from openpyxl import load_workbook
import sdek
import boxberry
import csv
from tqdm import tqdm
from dbf_light import Dbf

wb = load_workbook('82x465_24.08.xlsx')

sheet = wb['Лист1 (2)']

tariffs = {
    5: "Экономичный экспресс",
    10: "Экспресс лайт"
}

receivers_kladr = [sheet.cell(row=i, column=9).value for i in range(2, 38132)]
senders_kladr = [sheet.cell(row=i, column=5).value for i in range(2, 38132)]

receivers_fias = [sheet.cell(row=i, column=4).value for i in range(2, 38132)]
senders_fias = [sheet.cell(row=i, column=8).value for i in range(2, 38132)]

cities: Dict[str, str] = {}

with Dbf.open('KLADR.DBF') as dbf:
    for row in dbf:
        cities.setdefault(row[2], row[0])


with open('sdek.csv', mode='w', encoding='utf-8') as sdek_csv:
    file_writer = csv.writer(sdek_csv, delimiter=",", lineterminator="\r")
    file_writer.writerow(
        ["Ключ (ОткудаФИАС+КудаФИАС)", "метод доставки", "Цена", "Минимальный срок", "Максимальный срок"])
    for i in tqdm(range(len(senders_kladr))):
        sender_name = cities[senders_kladr[i]]
        receiver_name = cities[receivers_kladr[i]]
        sender_fias = senders_fias[i]
        receiver_fias = receivers_fias[i]
        try:
            response = sdek.get_sdek(sender_name, receiver_name)
            for tariff in response['result']:
                if tariff['status']:
                    file_writer.writerow([
                        f'{sender_fias}  {receiver_fias}',
                        f'{tariffs[tariff["tariffId"]]}',
                        f'{tariff["result"]["price"]}',
                        f'{tariff["result"]["deliveryPeriodMin"]}',
                        f'{tariff["result"]["deliveryPeriodMax"]}'
                    ])
                else:
                    file_writer.writerow([
                        f'{sender_fias}  {receiver_fias}',
                        f'{tariffs[tariff["tariffId"]]}',
                        "Невозможно осуществить доставку по этому направлению при заданных условиях"
                    ])
        except:
            file_writer.writerow(["Внутренняя ошибка сервера"])

with open("Boxberry.csv", mode='w', encoding='utf-8') as bb_csv:
    file_writer = csv.writer(bb_csv, delimiter=",", lineterminator="\r")
    file_writer.writerow(["Ключ (ОткудаФИАС+КудаФИАС)", "метод доставки", "Цена", "срок"])
    for i in tqdm(range(len(senders_kladr))):
        sender_name = cities[senders_kladr[i]]
        receiver_name = cities[receivers_kladr[i]]
        sender_fias = senders_fias[i]
        receiver_fias = receivers_fias[i]
        sender_city_id = boxberry.get_city_id(senders_kladr[i])
        receiver_city_id = boxberry.get_city_id(receivers_kladr[i])
        try:
            sender_point, receiver_point = boxberry.get_delivery_points(sender_city_id, receiver_city_id)
            delivery = boxberry.calculate_delivery(sender_point, receiver_point)
            file_writer.writerow([
                f'{sender_fias}  {receiver_fias}',
                'Склад-склад',
                f'{delivery["price"]}',
                f'{delivery["delivery_period"]}'
            ])
        except KeyError:
            file_writer.writerow([
                f'{sender_fias}  {receiver_fias}',
                'Доставка невозможна.'
            ])
